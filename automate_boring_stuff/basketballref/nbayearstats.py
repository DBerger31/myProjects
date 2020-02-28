#! /usr/bin/python3.7
# Webscraping NBA stats
# create2020stats.py - Asks user to input a year and will create an
# excel spreadsheet with the stats of all players during that year

import openpyxl
from urllib.request import urlopen
from bs4 import BeautifulSoup

class NBA:
    def getHeader(self):
        headers = [] # list for storing headers
        # use getText() to extract the text we need into a list
        for text in self.soup.findAll('tr')[0].findAll('th'):
            headers.append(text.getText())
        headers = headers[1:]
        return headers

    def getStats(self):
        rowsOfPlayers = self.soup.findAll('tr')[1:]
        # creates a 2d list of all the players and their stats
        player_stats = [[stats.getText() for stats in rowsOfPlayers[i].findAll('td')] for i in range(len(rowsOfPlayers))]
        return player_stats

    def statExcel(self):
        year = input("Enter what year: ")
        # URL page we will be scraping
        url = 'https://www.basketball-reference.com/leagues/NBA_{}_per_game.html'.format(year)
        # this is the HTML form the given url
        html = urlopen(url)
        # stores source html code in soup
        self.soup = BeautifulSoup(html, 'html.parser')
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'PerGame' + str(year)

        i = 1
        for header in self.getHeader():
            sheet.cell(row=1, column=i).value = header
            i += 1

        stats_2d_list = self.getStats()
        x = 1
        for r in stats_2d_list:
            y = 1
            x += 1
            for item in r:
                sheet.cell(row= x, column= y).value = item
                y += 1


        wb.save('nba{}to{}.xlsx'.format(str(int(year) - 1), year))

nba_work = NBA()
nba_work.statExcel()
